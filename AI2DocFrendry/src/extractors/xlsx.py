"""Excel ブック (.xlsx/.xlsm) から構造化ブロックを抽出する。

結合セルを展開して矩形の行列を保つことを最優先とする。行列の対応が
崩れた表は、渡された時点で元の文書とは別の内容になってしまうため。
"""

from __future__ import annotations

import datetime as _dt
from typing import TYPE_CHECKING, Any, Iterable, NamedTuple

from ..models import Block, ImageRef, SourceFile, StyleMark
from .base import ExtractionError, build

if TYPE_CHECKING:  # 実行時 import を避け、循環・未定義シンボルで落ちないようにする
    from ..models import RawDocument

# LLM は表を縦方向に読むのが苦手で、50行×5列で正答率0.81、
# 5行×50列では0.43まで落ちる。横長の表は事前に警告する。
MAX_COLUMNS = 20

_WIDE_TABLE_REASON = (
    "LLMは横長の表の読解精度が大きく落ちる（5行×50列で正答率0.43）。"
    "列を分割するか転置を検討してください。"
)


class _SheetParts(NamedTuple):
    """1シートから抽出した成果物。"""

    blocks: tuple[Block, ...]
    images: tuple[ImageRef, ...]
    style_marks: tuple[StyleMark, ...]
    warnings: tuple[str, ...]


class XlsxExtractor:
    """openpyxl を用いた Excel 抽出器。"""

    extensions: tuple[str, ...] = (".xlsx", ".xlsm")

    def extract(self, source: SourceFile) -> RawDocument:
        """ブックを読み取り RawDocument を返す。失敗は ExtractionError に包む。"""
        try:
            return self._extract(source)
        except ExtractionError:
            raise
        except Exception as exc:  # noqa: BLE001 - 契約上すべて ExtractionError に統一
            raise ExtractionError(f"xlsx の抽出に失敗しました: {source.path}: {exc}") from exc

    def _extract(self, source: SourceFile) -> RawDocument:
        """全シートを走査して成果物を組み立てる。"""
        values_wb, formulas_wb = _load_workbooks(source.path)
        try:
            parts = tuple(
                _extract_sheet(name, values_wb[name], formulas_wb[name])
                for name in values_wb.sheetnames
            )
        finally:
            _close_all((values_wb, formulas_wb))
        return build(
            source,
            tuple(block for part in parts for block in part.blocks),
            images=tuple(image for part in parts for image in part.images),
            style_marks=tuple(mark for part in parts for mark in part.style_marks),
            warnings=tuple(warning for part in parts for warning in part.warnings),
        )


def _load_workbooks(path: Any) -> tuple[Any, Any]:
    """計算結果値用と数式文字列用の 2 つのブックを開く。"""
    try:
        import openpyxl
    except ImportError as exc:
        raise ExtractionError("openpyxl がインストールされていません") from exc
    try:
        values_wb = openpyxl.load_workbook(str(path), data_only=True)
        formulas_wb = openpyxl.load_workbook(str(path), data_only=False)
    except Exception as exc:  # noqa: BLE001
        raise ExtractionError(f"ブックを開けませんでした: {path}: {exc}") from exc
    return values_wb, formulas_wb


def _close_all(workbooks: Iterable[Any]) -> None:
    """ブックを閉じる。閉じる際の失敗は抽出結果に影響しないため無視する。"""
    for workbook in workbooks:
        try:
            workbook.close()
        except Exception:  # noqa: BLE001,S110
            pass


def _extract_sheet(name: str, values_ws: Any, formulas_ws: Any) -> _SheetParts:
    """1シート分の見出し・表・画像・スタイルを抽出する。"""
    grid, used_formula = _sheet_grid(values_ws, formulas_ws)
    merged = _apply_merged(grid, _merged_ranges(values_ws))
    trimmed = _trim_trailing(merged)
    heading = Block(kind="heading", text=name, level=1, position=f"{name}!A1")
    blocks: tuple[Block, ...] = (heading,)
    if trimmed:
        blocks += (Block(kind="table", table_rows=_freeze(trimmed), position=f"{name}!A1"),)
    return _SheetParts(
        blocks=blocks,
        images=_sheet_images(name, values_ws),
        style_marks=_strike_marks(name, formulas_ws),
        warnings=_sheet_warnings(name, trimmed, used_formula),
    )


def _sheet_warnings(name: str, grid: list[list[str]], used_formula: bool) -> tuple[str, ...]:
    """列数超過と数式フォールバックの警告を組み立てる。"""
    warnings: tuple[str, ...] = ()
    width = max((len(row) for row in grid), default=0)
    if width > MAX_COLUMNS:
        warnings += (f"シート'{name}'の表が{width}列（{MAX_COLUMNS}列超）です。{_WIDE_TABLE_REASON}",)
    if used_formula:
        warnings += (
            f"シート'{name}'に計算結果値を持たないセルがあり、数式文字列で代替しました。",
        )
    return warnings


def _sheet_grid(values_ws: Any, formulas_ws: Any) -> tuple[list[list[str]], bool]:
    """セル値を文字列の二次元リストにする。値が無い数式は数式文字列で代替する。"""
    max_row = int(values_ws.max_row or 0)
    max_col = int(values_ws.max_column or 0)
    grid: list[list[str]] = []
    used_formula = False
    for row_index in range(1, max_row + 1):
        row: list[str] = []
        for col_index in range(1, max_col + 1):
            text = _cell_text(values_ws.cell(row=row_index, column=col_index).value)
            if not text:
                formula = _formula_text(formulas_ws, row_index, col_index)
                if formula:
                    text = formula
                    used_formula = True
            row.append(text)
        grid.append(row)
    return grid, used_formula


def _formula_text(formulas_ws: Any, row_index: int, col_index: int) -> str:
    """数式セルであれば数式文字列を返す。"""
    try:
        raw = formulas_ws.cell(row=row_index, column=col_index).value
    except Exception:  # noqa: BLE001
        return ""
    return raw if isinstance(raw, str) and raw.startswith("=") else ""


def _cell_text(value: Any) -> str:
    """セル値を表示用の文字列に正規化する。"""
    if value is None:
        return ""
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    if isinstance(value, (_dt.datetime, _dt.date, _dt.time)):
        return value.isoformat()
    return str(value).strip()


def _merged_ranges(worksheet: Any) -> tuple[tuple[int, int, int, int], ...]:
    """結合セル範囲を (min_row, min_col, max_row, max_col) の組で返す。"""
    ranges = getattr(getattr(worksheet, "merged_cells", None), "ranges", None) or ()
    result: tuple[tuple[int, int, int, int], ...] = ()
    for cell_range in ranges:
        try:
            result += (
                (
                    int(cell_range.min_row),
                    int(cell_range.min_col),
                    int(cell_range.max_row),
                    int(cell_range.max_col),
                ),
            )
        except (AttributeError, TypeError, ValueError):
            continue
    return result


def _apply_merged(
    grid: list[list[str]], ranges: tuple[tuple[int, int, int, int], ...]
) -> list[list[str]]:
    """結合範囲の値を範囲内の全セルへ複製し、矩形の行列対応を保つ。"""
    if not grid or not ranges:
        return [list(row) for row in grid]
    expanded = [list(row) for row in grid]
    for min_row, min_col, max_row, max_col in ranges:
        if min_row - 1 >= len(expanded) or min_col - 1 >= len(expanded[min_row - 1]):
            continue
        value = expanded[min_row - 1][min_col - 1]
        for row_index in range(min_row - 1, min(max_row, len(expanded))):
            row = expanded[row_index]
            for col_index in range(min_col - 1, min(max_col, len(row))):
                row[col_index] = value
    return expanded


def _trim_trailing(grid: list[list[str]]) -> list[list[str]]:
    """末尾の空行・空列のみを削る。中間の空行は構造として残す。"""
    rows = [list(row) for row in grid]
    while rows and not any(cell for cell in rows[-1]):
        rows.pop()
    if not rows:
        return []
    last_col = 0
    for row in rows:
        for index, cell in enumerate(row):
            if cell:
                last_col = max(last_col, index + 1)
    if last_col == 0:
        return []
    return [row[:last_col] for row in rows]


def _freeze(grid: list[list[str]]) -> tuple[tuple[str, ...], ...]:
    """行長を揃えた不変の行列にする。

    base.expand_merged は空セルを直前の値で埋めるため、ここでは使わない。
    結合範囲は _apply_merged で厳密に展開済みで、残る空セルは本当に空である。
    """
    if not grid:
        return ()
    width = max(len(row) for row in grid)
    return tuple(tuple(row) + ("",) * (width - len(row)) for row in grid)


def _strike_marks(name: str, worksheet: Any) -> tuple[StyleMark, ...]:
    """打ち消し線付きセルを StyleMark として収集する。"""
    marks: tuple[StyleMark, ...] = ()
    max_row = int(worksheet.max_row or 0)
    max_col = int(worksheet.max_column or 0)
    for row_index in range(1, max_row + 1):
        for col_index in range(1, max_col + 1):
            cell = worksheet.cell(row=row_index, column=col_index)
            if not getattr(getattr(cell, "font", None), "strike", False):
                continue
            text = _cell_text(cell.value)
            if not text:
                continue
            position = f"{name}!{_column_letter(col_index)}{row_index}"
            marks += (
                StyleMark(kind="strikethrough", text=text, position=position, preserved=False),
            )
    return marks


def _sheet_images(name: str, worksheet: Any) -> tuple[ImageRef, ...]:
    """埋め込み画像とグラフを ImageRef として収集する。"""
    refs: tuple[ImageRef, ...] = ()
    for image in getattr(worksheet, "_images", None) or ():
        refs += (ImageRef(position=_anchor_position(name, image), count=1, kind="image"),)
    for chart in getattr(worksheet, "_charts", None) or ():
        refs += (ImageRef(position=_anchor_position(name, chart), count=1, kind="chart"),)
    return refs


def _anchor_position(name: str, drawing: Any) -> str:
    """描画オブジェクトのアンカー位置を "シート名!セル番地" に変換する。"""
    anchor = getattr(drawing, "anchor", None)
    if isinstance(anchor, str):
        return f"{name}!{anchor}"
    origin = getattr(anchor, "_from", None)
    if origin is None:
        return f"{name}!A1"
    try:
        return f"{name}!{_column_letter(int(origin.col) + 1)}{int(origin.row) + 1}"
    except (AttributeError, TypeError, ValueError):
        return f"{name}!A1"


def _column_letter(index: int) -> str:
    """1 始まりの列番号を A, B, ..., AA 形式に変換する。"""
    if index < 1:
        return "A"
    letters = ""
    remaining = index
    while remaining > 0:
        remaining, offset = divmod(remaining - 1, 26)
        letters = chr(ord("A") + offset) + letters
    return letters
